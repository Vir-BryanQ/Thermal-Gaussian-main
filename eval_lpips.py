import os
import gc
import sys
import torch
import random
import json
import shutil
import subprocess
import argparse
import pandas as pd
import signal
import time
from datetime import datetime
import psutil
import uuid
from tqdm import tqdm
from utils.general_utils import safe_state
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import deque

def parse_arguments():
    parser = argparse.ArgumentParser(description='Training and evaluation loop')
    parser.add_argument('--dataset_path', type=str, required=True, help='Path to the dataset')
    parser.add_argument('--output_excel', type=str, required=True, help='Name of the output Excel file')
    parser.add_argument('--date_lists', type=str, required=True, help='Path to the dataset')
    return parser.parse_args()

def append_row(output_excel, row_values):
    wb = load_workbook(output_excel)
    ws = wb.active
    ws.append(row_values)
    wb.save(output_excel)


def perform_sampling(args):
    if os.path.exists(args.output_excel):
        print(f'Deleting {args.output_excel}.')
        os.remove(args.output_excel)  
    wb = Workbook()
    ws = wb.active
    ws.append([
       'lpips rgb', 'lpips th', 'date', 
    ])
    bold_font = Font(bold=True)
    for cell in ws[1]: 
        cell.font = bold_font
    wb.save(args.output_excel)


    date_lists_path = os.path.join(args.dataset_path, args.date_lists)
    with open(date_lists_path, 'r') as f:
        date_lists = [line.strip() for line in f.read().splitlines() if line.strip()]

    dataset_name = os.path.basename(args.dataset_path.rstrip('/'))
    thgs_dir = os.path.join("outputs", dataset_name)

    for i in tqdm(range(len(date_lists))):
        date_file = date_lists[i]

        model_path = os.path.join(thgs_dir, date_file)

        eval_command = f"python metrics.py --model_paths {model_path} --ns_metric"
        print(eval_command)
        process = subprocess.Popen(eval_command, shell=True, env=os.environ)
        process.wait()

        eval_json_path = os.path.join(model_path, 'results.json')
        with open(eval_json_path, 'r') as f:
            eval_results = json.load(f)

        lpips_rgb = eval_results['ours_30000']['color_LPIPS']
        lpips_th = eval_results['ours_30000']['thermal_LPIPS']

        append_row(args.output_excel, [
            lpips_rgb , lpips_th, date_file
            ])

        print(f'Row-{i} in {args.output_excel} saved.')


if __name__ == '__main__':
    args = parse_arguments()
    safe_state(False)
    perform_sampling(args)
