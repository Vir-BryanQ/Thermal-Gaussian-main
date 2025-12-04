import os
import gc
import sys
import torch
import random
import json
import shutil
import subprocess
import argparse
import pandas as pd
import signal
import time
from datetime import datetime
import psutil
import uuid
from tqdm import tqdm
from utils.general_utils import safe_state
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import deque

def preallocate_vmem(vram=38):
    required_elements = int(vram * 1024 * 1024 * 1024 / 4)
    while True:
        try:
            occupied = torch.empty(required_elements , dtype=torch.float32, device='cuda')
            del occupied
            break
        except RuntimeError as e:
            t = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            if 'CUDA out of memory' in str(e):
                free_memory, total_memory = torch.cuda.mem_get_info(torch.cuda.current_device())
                print(f"*** CUDA OOM: {free_memory / (1024 * 1024)}MiB is free [{t}]")
            else:
                print(f"### {e} [{t}]")

def release_vmem():
    torch.cuda.empty_cache()

def parse_arguments():
    parser = argparse.ArgumentParser(description='Training and evaluation loop')
    parser.add_argument('--dataset_path', type=str, required=True, help='Path to the dataset')
    parser.add_argument('--output_excel', type=str, required=True, help='Name of the output Excel file')
    parser.add_argument('--train_lists', type=str, required=True, help='Path to the dataset')
    parser.add_argument('--iter', action='store_true')
    return parser.parse_args()

def append_row(output_excel, row_values):
    wb = load_workbook(output_excel)
    ws = wb.active
    ws.append(row_values)
    wb.save(output_excel)


def perform_sampling(args):
    preallocate_vmem()

    if os.path.exists(args.output_excel):
        print(f'Deleting {args.output_excel}.')
        os.remove(args.output_excel)  
    wb = Workbook()
    ws = wb.active
    ws.append([
        'psnr rgb', 'ssim rgb', 'lpips rgb', 'psnr th', 'ssim th', 'lpips th', 'iter', 'time', 'date', 'file', 
    ])
    bold_font = Font(bold=True)
    for cell in ws[1]: 
        cell.font = bold_font
    wb.save(args.output_excel)


    train_lists_path = os.path.join(args.dataset_path, args.train_lists)
    with open(train_lists_path, 'r') as f:
        train_lists = [line.strip() for line in f.read().splitlines() if line.strip()]

    dataset_name = os.path.basename(args.dataset_path.rstrip('/'))
    thgs_dir = os.path.join("outputs", dataset_name)
    os.makedirs(thgs_dir, exist_ok=True)

    for i in tqdm(range(len(train_lists))):
        train_file = train_lists[i]

        if args.iter:
            train_file_path = os.path.join(args.dataset_path, train_file)
            with open(train_file_path, 'r') as f:
                training_imgs = [line.strip() for line in f.read().splitlines() if line.strip()]
            training_iterations = len(training_imgs) * 50
        else:
            training_iterations = 30000

        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        model_path = os.path.join(thgs_dir, ts)

        training_command = (f"python train-OMMG.py -s {args.dataset_path} -m {model_path} --train_file {train_file} --iterations {training_iterations} --no_report")
        print(training_command)
        process = subprocess.Popen(training_command, shell=True, env=os.environ, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, bufsize=1)

        time.sleep(1)
        release_vmem()

        training_time = None
        for line in process.stdout:
            # print("[child]", line, end="")
            training_time = line.strip()  
        print(training_time)
        training_time = f"{float(training_time):.1f}s"

        process.wait()

        preallocate_vmem()

        render_command = (f"python render.py --model_path {model_path} --skip_train")
        print(render_command)
        process = subprocess.Popen(render_command, shell=True, env=os.environ)

        time.sleep(1)
        release_vmem()

        process.wait()
        preallocate_vmem()

        eval_command = f"python metrics.py --model_paths {model_path} --ns_metric"
        print(eval_command)
        process = subprocess.Popen(eval_command, shell=True, env=os.environ)

        time.sleep(1)
        release_vmem()
        process.wait()

        preallocate_vmem()

        eval_json_path = os.path.join(model_path, 'results.json')
        with open(eval_json_path, 'r') as f:
            eval_results = json.load(f)

        ssim_rgb = eval_results[f'ours_{training_iterations}']['color_SSIM']
        psnr_rgb = eval_results[f'ours_{training_iterations}']['color_PSNR']
        lpips_rgb = eval_results[f'ours_{training_iterations}']['color_LPIPS']
        ssim_th = eval_results[f'ours_{training_iterations}']['thermal_SSIM']
        psnr_th = eval_results[f'ours_{training_iterations}']['thermal_PSNR']
        lpips_th = eval_results[f'ours_{training_iterations}']['thermal_LPIPS']

        append_row(args.output_excel, [
            psnr_rgb, ssim_rgb, lpips_rgb , psnr_th, ssim_th, lpips_th, training_iterations, training_time, ts, train_file
            ])

        print(f'Row-{i} in {args.output_excel} saved.')


if __name__ == '__main__':
    args = parse_arguments()
    safe_state(False)
    perform_sampling(args)
